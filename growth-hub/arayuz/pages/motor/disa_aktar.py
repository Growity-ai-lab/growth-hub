# -*- coding: utf-8 -*-
"""
T&G Growth Hub — plan dışa aktarma (Excel).

Karar katmanının bir çıktısı: seçilen planı sade bir Excel'e döker. Şablon (v2) sonra;
bu sürüm okunur bir çalışma dosyası üretir. Puanı/birim maliyeti YENİDEN HESAPLAMAZ (K3) —
kayıttaki değerleri yazar; tahmini sonuç yalnız görüntü amaçlı (bütçe ÷ birim).

plan_excel_b64(kayit) -> base64 (xlsx). kayit = arayüzün 'secim' kaydı (brief + satirlar).
Tarayıcıda (Pyodide/openpyxl) ya da yerel sunucuda aynı kodla çalışır.
"""
import io, base64


def _hacim(butce, birim, tip):
    if not birim or birim <= 0:
        return None
    return butce / birim * 1000 if tip == 'CPM' else butce / birim


_METRIK = {'CPM': 'gösterim', 'CPV': 'izlenme', 'CPC': 'tıklama'}


def plan_excel_b64(kayit):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    b = kayit.get('brief', {}) or {}
    satirlar = kayit.get('satirlar', []) or []
    kpi = b.get('kpi_projeksiyon', {}) or {}
    hedef = b.get('kpi_hedef', {}) or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Plan'
    kalin = Font(name='Arial', bold=True)
    baslik = Font(name='Arial', bold=True, color='FFFFFF')
    dolgu = PatternFill('solid', fgColor='1F3864')
    ort = Alignment(horizontal='center')
    ince = Side(style='thin', color='BFBFBF')
    kutu = Border(ince, ince, ince, ince)

    ws['A1'] = 'Growth Hub — Mecra Planı'
    ws['A1'].font = Font(name='Arial', bold=True, size=14)

    def kv(satir, etiket, deger):
        ws.cell(satir, 1, etiket).font = kalin
        ws.cell(satir, 2, deger)

    kv(3, 'Marka', b.get('marka') or '—')
    kv(4, 'Ürün', b.get('urun') or '—')
    kv(5, 'Sektör', b.get('sektor_l2') or '—')
    kv(6, 'Amaç', b.get('amac') or '—')
    kv(7, 'Toplam bütçe (₺)', b.get('toplam_butce') or 0)
    kv(8, 'Tarih', kayit.get('zaman') or '')
    kv(9, 'Tahmini sonuç',
       f"~{round(kpi.get('gosterim', 0)):,} gösterim · ~{round(kpi.get('izlenme', 0)):,} izlenme · ~{round(kpi.get('tiklama', 0)):,} tıklanma".replace(',', '.'))

    b0 = 11
    kolonlar = ['Yayıncı', 'Grup', 'Reklam modeli', 'Fiyat tipi', 'Önerilen birim',
                'Bütçe (₺)', 'Tahmini sonuç', 'Sistem bütçesi', 'Sapma (₺)', 'Kaynak', 'Sebep']
    for j, h in enumerate(kolonlar, 1):
        c = ws.cell(b0, j, h)
        c.font = baslik
        c.fill = dolgu
        c.alignment = ort
        c.border = kutu

    r = b0 + 1
    for s in satirlar:
        birim = s.get('secilen_birim') or s.get('sistem_birim')
        butce = s.get('secilen_butce') or 0
        v = _hacim(butce, birim, s.get('tip'))
        sonuc = f"~{round(v):,} {_METRIK.get(s.get('tip'), 'birim')}".replace(',', '.') if v is not None else '—'
        deger = [s.get('yayinci'), s.get('grup'), s.get('reklam_modeli'), s.get('tip'),
                 birim, butce, sonuc, s.get('sistem_butce'),
                 s.get('sapma_butce'), s.get('kaynak') or ('öneri' if s.get('sistem_butce') else 'ekip'),
                 s.get('sebep') or '']
        for j, d in enumerate(deger, 1):
            c = ws.cell(r, j, d)
            c.border = kutu
            if j in (5,):
                c.number_format = '#,##0.0000'
            if j in (6, 8, 9):
                c.number_format = '#,##0'
        r += 1

    # toplam satırı
    ws.cell(r, 5, 'Toplam').font = kalin
    tc = ws.cell(r, 6, sum((s.get('secilen_butce') or 0) for s in satirlar))
    tc.font = kalin
    tc.number_format = '#,##0'

    genis = [16, 14, 18, 9, 13, 13, 22, 14, 12, 8, 30]
    for j, g in enumerate(genis, 1):
        ws.column_dimensions[get_column_letter(j)].width = g

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode('ascii')
