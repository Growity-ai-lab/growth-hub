# -*- coding: utf-8 -*-
"""
T&G Growth Hub — Puanlama motoru.

Tek işi var: dönüştürülmüş kapanış raporlarından mecra karnesini üretmek.
Hiçbir eşik, katsayı ya da sınır bu dosyada gömülü DEĞİLDİR; hepsi
../sozluk/sozluk.xlsx içindeki Parametreler sayfasından okunur.

Kullanım:  python3 skor.py <donusum.json> [cikti_klasoru]
"""
import json, math, os, sys, statistics, collections, datetime

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'sozluk'))
from sozluk import Sozluk

SOZLUK = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'sozluk', 'sozluk.xlsx')


def satirlari_hazirla(kampanyalar, bugun):
    """Her yayınlanmış satır için etkin birim maliyeti ve zaman ağırlığını hesaplar."""
    satirlar = []
    for k in kampanyalar:
        for x in k['satirlar']:
            if x['durum'] != 'Yayınlandı' or not x['ger_hacim'] or not x['ger_bedel']:
                continue
            carpan = 1000 if x['tip'] == 'CPM' else 1
            satirlar.append(dict(
                s2=k['s2'], grup=x['grup'], yayinci=x['mecra'],
                tur=x.get('ana_tur') or 'Diğer',
                reklam_modeli=(x.get('reklam_modeli') or x.get('ana_tur') or 'Diğer'),
                ham_tur=x['format'], tip=x['tip'],
                birim=x['ger_bedel'] / x['ger_hacim'] * carpan,
                bedel=x['ger_bedel'],
                plan_birim=(x['plan_bedel'] / x['plan_hacim'] * carpan)
                           if x.get('plan_bedel') and x.get('plan_hacim') else (x.get('plan_birim') or None),
                vcr=x.get('vcr'), ahb=x.get('ahb'),
                teslim=min(x['ger_hacim'] / x['plan_hacim'], 3.0) if x['plan_hacim'] else None,
                bas=k['bas'], kampanya=k['dosya']))
    return satirlar


def puanla(satirlar, sz, bugun=None):
    bugun = bugun or datetime.date.today()
    havuz_min = int(sz.p('peer_havuzu_min', 3))
    yari_omur = float(sz.p('zaman_yari_omru_ay', 18))
    k_cekme = float(sz.p('cekme_katsayisi', 6))
    guven = int(sz.p('guven_esigi', 8))
    # kırılım seviyesi ajans politikası (sözlük). "ana tür" = bugünkü davranış (geri uyum).
    fine = str(sz.p('karne_kirilim', 'ana tür')).strip().lower() == 'reklam modeli'
    for r in satirlar:                     # bu satırın karne kırılımı: kapalıysa ana tür'e çöker
        r['rm'] = r['reklam_modeli'] if fine else r['tur']

    # karşılaştırma havuzu — geri çekilmeli: reklam modeli → ana tür → tüm sektörler
    rm_pool, dar, genis = (collections.defaultdict(list) for _ in range(3))
    for r in satirlar:
        if fine:
            rm_pool[(r['s2'], r['tip'], r['rm'])].append(r['birim'])
        dar[(r['s2'], r['tip'], r['tur'])].append(r['birim'])
        genis[(r['tip'], r['tur'])].append(r['birim'])

    for r in satirlar:
        if fine and len(rm_pool[(r['s2'], r['tip'], r['rm'])]) >= havuz_min:
            ref = rm_pool[(r['s2'], r['tip'], r['rm'])]; seviye = 'sektör + reklam modeli'
        elif len(dar[(r['s2'], r['tip'], r['tur'])]) >= havuz_min:
            ref = dar[(r['s2'], r['tip'], r['tur'])]; seviye = 'sektör + yayın türü'
        else:
            ref = genis[(r['tip'], r['tur'])]; seviye = 'tüm sektörler, yayın türü'
        orta = statistics.median(ref) if ref else r['birim']
        r['indeks'] = max(20, min(220, orta / r['birim'] * 100)) if r['birim'] else 100
        r['seviye'] = seviye
        ay = (bugun - datetime.date.fromisoformat(r['bas'])).days / 30.4
        r['agirlik'] = math.log(1 + r['bedel']) * (0.5 ** (ay / yari_omur))

    # önerilen birim maliyet için geri çekilme havuzları (reklam modeli → yayıncı → grup)
    p_rm, p_tam, p_yay, p_grup = (collections.defaultdict(list) for _ in range(4))
    for r in satirlar:
        if fine:
            p_rm[(r['s2'], r['yayinci'], r['rm'], r['tip'])].append(r['birim'])
        p_tam[(r['s2'], r['yayinci'], r['tur'], r['tip'])].append(r['birim'])
        p_yay[(r['yayinci'], r['tur'], r['tip'])].append(r['birim'])
        p_grup[(r['grup'], r['tur'], r['tip'])].append(r['birim'])

    hucreler = collections.defaultdict(list)
    for r in satirlar:
        hucreler[(r['s2'], r['grup'], r['yayinci'], r['tur'], r['rm'], r['tip'])].append(r)

    sonuc = []
    for (s2, grup, yayinci, tur, rm, tip), rs in hucreler.items():
        W = sum(r['agirlik'] for r in rs); n = len(rs)
        ham = sum(r['indeks'] * r['agirlik'] for r in rs) / W
        birim_agirlik = W / n
        puan = (ham * W + 100 * k_cekme * birim_agirlik) / (W + k_cekme * birim_agirlik)
        teslimler = [r['teslim'] for r in rs if r['teslim']]

        merdiven = []
        if fine:
            merdiven.append(((s2, yayinci, rm, tip), p_rm, 'bu yayıncı, bu reklam modeli, bu sektör'))
        merdiven += [((s2, yayinci, tur, tip), p_tam, 'bu yayıncı, bu sektör'),
                     ((yayinci, tur, tip), p_yay, 'bu yayıncı, tüm sektörler'),
                     ((grup, tur, tip), p_grup, 'yayıncı grubu ortalaması')]
        for anahtar, havuz, etiket in merdiven:
            v = havuz.get(anahtar, [])
            if v:
                oner, oner_kaynak, oner_n = statistics.median(v), etiket, len(v); break
        else:
            oner, oner_kaynak, oner_n = statistics.median([r['birim'] for r in rs]), 'tek gözlem', n

        # Fayda = Puan × Yayınlanma% (teslim ile düzeltilmiş verimlilik). Teslim yoksa nötr (1.0).
        teslim_cell = round(statistics.mean(teslimler) * 100) if teslimler else None
        teslim_orani = (teslim_cell / 100) if teslim_cell is not None else 1.0
        fayda = round(round(puan, 1) * teslim_orani, 1)

        # Planlanan vs gerçekleşen kanıtı (seçim gezgininde "puan ↔ sonuçlar" geçişi için).
        plan_birimler = [r['plan_birim'] for r in rs if r.get('plan_birim')]
        vcrler = [r['vcr'] for r in rs if r.get('vcr') is not None]
        ahblar = [r['ahb'] for r in rs if r.get('ahb') is not None]
        ger_birim_cell = round(statistics.median([r['birim'] for r in rs]), 4)
        plan_birim_cell = round(statistics.median(plan_birimler), 4) if plan_birimler else None
        birim_sapma = round((ger_birim_cell / plan_birim_cell - 1) * 100) if plan_birim_cell else None

        sonuc.append(dict(
            s2=s2, grup=grup, yayinci=yayinci, tur=tur, reklam_modeli=rm, tip=tip, n=n,
            puan=round(puan, 1), fayda=fayda, aralik=round(14 + 62 / math.sqrt(n), 1),
            teslim=teslim_cell,
            plan_birim=plan_birim_cell, ger_birim=ger_birim_cell, birim_sapma=birim_sapma,
            vcr=(round(statistics.mean(vcrler) * 100, 1) if vcrler else None),
            ahb=(round(statistics.mean(ahblar), 2) if ahblar else None),
            harcama=round(sum(r['bedel'] for r in rs)),
            kampanya=len({r['kampanya'] for r in rs}),
            oner=round(oner, 4), oner_kaynak=oner_kaynak, oner_n=oner_n,
            ornek_tur=collections.Counter(r['ham_tur'] for r in rs).most_common(1)[0][0],
            son=max(r['bas'] for r in rs)[:7],
            guven=('kendi verisi' if n >= guven else
                   ('yarısı tahmin' if n >= guven / 2 else 'tahmin')),
            havuz=rs[0]['seviye'],
            kirilim_seviye=('reklam modeli' if fine else 'ana tür')))
    sonuc.sort(key=lambda x: (x['s2'], x['tur'], -x['puan']))
    return sonuc


def yaz_excel(sonuc, yol):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Mecra Karnesi'
    F = 'Arial'
    HF = Font(name=F, size=9, bold=True, color='FFFFFF'); HB = PatternFill('solid', fgColor='1F3864')
    BK = Font(name=F, size=10); t = Side(style='thin', color='BFBFBF')
    BOX = Border(t, t, t, t)
    bas = ['Sektör L2', 'Yayıncı Grubu', 'Yayıncı', 'Yayın Türü', 'Reklam modeli', 'Örnek yayın türü',
           'Fiyat Tipi', 'Puan', 'Fayda', 'Puan aralığı', 'Kaç satır', 'Kaç kampanya',
           'Ortalama yayınlanma %', 'Önerilen birim maliyet', 'Öneri neye dayanıyor', 'Kaç gözlemden',
           'Toplam harcama', 'Puan nereden geliyor', 'Karşılaştırma havuzu', 'Kırılım', 'Son kampanya']
    gen = [16, 15, 17, 20, 18, 22, 8, 8, 8, 12, 9, 12, 17, 15, 24, 12, 15, 18, 24, 14, 12]
    idx_oner = bas.index('Önerilen birim maliyet') + 1
    idx_harc = bas.index('Toplam harcama') + 1
    for j, h in enumerate(bas, 1):
        c = ws.cell(1, j, h); c.font = HF; c.fill = HB
        c.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[get_column_letter(j)].width = gen[j - 1]
    for i, x in enumerate(sonuc, 2):
        vals = [x['s2'], x['grup'], x['yayinci'], x['tur'], x['reklam_modeli'], x['ornek_tur'], x['tip'],
                x['puan'], x['fayda'],
                f"{round(x['puan']-x['aralik'])}–{round(x['puan']+x['aralik'])}", x['n'], x['kampanya'],
                x['teslim'], x['oner'], x['oner_kaynak'], x['oner_n'], x['harcama'],
                x['guven'], x['havuz'], x['kirilim_seviye'], x['son']]
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v); c.font = BK; c.border = BOX
            if j == idx_oner: c.number_format = '#,##0.0000'
            if j == idx_harc: c.number_format = '#,##0'
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(bas))}{len(sonuc)+1}'
    ws.sheet_view.showGridLines = False
    wb.save(yol)


if __name__ == '__main__':
    kaynak = sys.argv[1] if len(sys.argv) > 1 else '/tmp/donusum.json'
    cikti = sys.argv[2] if len(sys.argv) > 2 else '.'
    sz = Sozluk(SOZLUK)
    kampanyalar = json.load(open(kaynak, encoding='utf-8'))
    satirlar = satirlari_hazirla(kampanyalar, datetime.date.today())
    sonuc = puanla(satirlar, sz)
    json.dump(sonuc, open(os.path.join(cikti, 'karne.json'), 'w', encoding='utf-8'), ensure_ascii=False)
    yaz_excel(sonuc, os.path.join(cikti, 'Mecra_Karnesi.xlsx'))
    guven = int(sz.p('guven_esigi', 8))
    print(f'{len(satirlar)} yayınlanmış satır → {len(sonuc)} karne hücresi')
    print(f'güven eşiğini ({guven} satır) geçen hücre: '
          f"{sum(1 for x in sonuc if x['n'] >= guven)} / {len(sonuc)}")
    print('parametreler:', {k: sz.p(k) for k in
                            ['guven_esigi', 'zaman_yari_omru_ay', 'cekme_katsayisi',
                             'peer_havuzu_min', 'tek_grup_tavani', 'deneme_payi', 'uyari_esigi']})
