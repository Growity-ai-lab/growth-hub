# -*- coding: utf-8 -*-
"""
T&G Growth Hub — Eski gerçekleşen raporları standart şablona çevirir.

Yayıncı, yayın türü, sektör, fiyat tipi ve parametrelerin hiçbiri bu dosyada gömülü DEĞİLDİR;
hepsi ../sozluk/sozluk.xlsx dosyasından okunur. Yeni bir yayıncı ya da yazım hatası için
kodu değil, o Excel'i güncelleyin.

Kullanım:  python3 donustur.py <rapor_klasörü> [sozluk.xlsx yolu]
"""
import openpyxl, re, sys, os, json, datetime
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'sozluk'))
from sozluk import Sozluk, norm as _snorm

# ---------------------------------------------------------------- yardımcılar
def norm(s):
    if s is None: return ''
    s = str(s).replace('\n', ' ').replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)) and not isinstance(v, bool): return float(v)
    s = str(v).strip()
    if s in ('_', '-', '', 'N/A', '#DIV/0!', 'nan'): return None
    s = s.replace('%', '').replace('.', '').replace(',', '.') if re.match(r'^[\d.,]+$', s) else s
    try: return float(s)
    except Exception: return None

def txt(v):
    if v is None: return ''
    s = str(v).replace('\n', ' ').strip()
    return '' if s in ('_', '-') else re.sub(r'\s+', ' ', s)

# ---------------------------------------------------------------- sözlük (dosyadan)
SZ = None   # main() içinde yüklenir

AY = {'ocak':1,'şubat':2,'subat':2,'mart':3,'nisan':4,'mayıs':5,'mayis':5,'haziran':6,
      'temmuz':7,'ağustos':8,'agustos':8,'eylül':9,'eylul':9,'ekim':10,'kasım':11,'kasim':11,'aralık':12,'aralik':12}
SON = {1:31,2:28,3:31,4:30,5:31,6:30,7:31,8:31,9:30,10:31,11:30,12:31}

def donem_tarih(donem, yil):
    """'15 Mayıs- 15 Haziran' / 'Kasım' / '21 - 29 Mayıs' -> (başlangıç, bitiş, kesin_mi)"""
    d = norm(donem)
    aylar = [(m.start(), AY[m.group(0)]) for m in re.finditer('|'.join(AY.keys()), d)]
    if not aylar: return (None, None, False)
    gunler = [(m.start(), int(m.group(0))) for m in re.finditer(r'\b(\d{1,2})\b', d)]
    ilk_ay, son_ay = aylar[0][1], aylar[-1][1]
    g1 = g2 = None
    for poz, g in gunler:
        if poz < aylar[0][0] + 12 and g1 is None: g1 = g
        elif g2 is None: g2 = g
    kesin = bool(gunler)
    b = datetime.date(yil, ilk_ay, min(g1 or 1, SON[ilk_ay]))
    s = datetime.date(yil, son_ay, min(g2 or (g1 if ilk_ay == son_ay and g1 and not g2 else SON[son_ay]), SON[son_ay]))
    if s < b: s = datetime.date(yil, son_ay, SON[son_ay])
    bugun = datetime.date.today()
    if s > bugun:                      # gelecekteki tarih olamaz -> bir önceki yıl
        b = b.replace(year=b.year - 1); s = s.replace(year=s.year - 1)
    return (b.isoformat(), s.isoformat(), kesin)

def yil_bul(dosya, ws):
    m = re.search(r'20(2\d)', dosya)
    if m: return int(m.group(0))
    for r in range(1, 12):
        for c in range(1, 8):
            m = re.search(r'20(2\d)', str(ws.cell(r, c).value or ''))
            if m: return int(m.group(0))
    return 2026

# ---------------------------------------------------------------- fiyat tipi
def kaynak(plan_birim, ger_birim):
    """Gerçekleşen birim maliyet plandan farklıysa gerçek ölçüm vardır."""
    if plan_birim and ger_birim:
        if abs(ger_birim - plan_birim) / max(plan_birim, 1e-9) > 0.005:
            return 'Platform Raporu'
    return 'Hesaplanan'

# ---------------------------------------------------------------- okuma
def basliklar(ws, r, cmax):
    return {c: norm(ws.cell(r, c).value) for c in range(1, cmax + 1)}

def bul(hdr, *parcalar, tam=False):
    for c, h in hdr.items():
        if not h: continue
        if tam and h in parcalar: return c
        if not tam and all(p in h for p in parcalar): return c
    return None

def oku_dosya(yol):
    ad = os.path.basename(yol)
    wb = openpyxl.load_workbook(yol, data_only=True)
    ws = None
    for isim in wb.sheetnames:
        w = wb[isim]
        for r in range(1, min(25, w.max_row + 1)):
            if any(norm(w.cell(r, c).value) in ('platform', 'marka') and
                   norm(w.cell(r, c + 1).value) in ('mecra', 'kampanya')
                   for c in range(1, min(6, w.max_column))):
                ws = w; break
        if ws is not None: break
    if ws is None: return None

    # --- künye
    meta = {}
    for r in range(1, 14):
        for c in range(1, 5):
            et = norm(ws.cell(r, c).value)
            if et.endswith(':') or et in ('gerçekleşen stopaj', 'gerçekleşen iletişim vergisi'):
                for cc in range(c + 1, c + 6):
                    v = ws.cell(r, cc).value
                    if v not in (None, ''):
                        meta[et.rstrip(':')] = v; break
    marka = txt(meta.get('marka'))
    kampanya = txt(meta.get('kampanya adı'))
    donem = txt(meta.get('kampanya dönemi'))
    yil = yil_bul(ad, ws)
    b_tar, s_tar, kesin = donem_tarih(donem, yil)
    s1, s2, s3, sektor_yok = SZ.sektor(marka)

    butce = None
    for k in ('planlanan herşey dahil kampanya bütçesi', 'planlanan kampanya maliyeti'):
        if k in meta: butce = num(meta[k]); break

    satirlar, uyari = [], []
    hdr, kolonlar = None, {}
    for r in range(1, ws.max_row + 1):
        ilk = [norm(ws.cell(r, c).value) for c in range(1, 6)]
        if 'platform' in ilk or ('marka' in ilk and 'kampanya' in ilk):
            hdr = basliklar(ws, r, ws.max_column)
            kolonlar = {
                'cihaz':   bul(hdr, 'platform', tam=True),
                'mecra':   bul(hdr, 'mecra', tam=True),
                'site':    bul(hdr, 'site/network'),
                'hedef':   bul(hdr, 'kategori', 'hedefleme'),
                'format':  bul(hdr, 'yayın türü'),
                'aset':    bul(hdr, 'aset', tam=True),
                'plan_h':  bul(hdr, 'planlanan imp'),
                'plan_b':  bul(hdr, 'planlanan unit cost') or bul(hdr, 'birim maliyet'),
                'plan_bd': bul(hdr, 'planlanan yayın bedeli') or bul(hdr, 'yayın bedeli', tam=True),
                'ger_bd':  bul(hdr, 'gerçekleşen yayın maliyeti') or bul(hdr, 'harcanan tutar'),
                'ahb':     bul(hdr, 'gerçekleşen ahb'),
                'adsrv':   bul(hdr, 'gerçekleşen adserver'),
                'reach':   bul(hdr, 'reach', tam=True),
                'imp':     bul(hdr, 'gerçekleşen impression') or bul(hdr, 'impression', tam=True),
                'view':    bul(hdr, 'gerçekleşen view') or bul(hdr, 'view', tam=True),
                'click':   bul(hdr, 'gerçekleşen click') or bul(hdr, 'click', tam=True),
                'vcr':     bul(hdr, 'vcr'),
                'ger_cpm': bul(hdr, 'gerçekleşen cpm'),
                'ger_cpv': bul(hdr, 'gerçekleşen cpv'),
                'tarih':   bul(hdr, 'yayın tarihleri'),
            }
            if kolonlar['plan_h']:
                kolonlar['ger_h'] = kolonlar['plan_h'] + 1 if norm(hdr.get(kolonlar['plan_h'] + 1, '')).startswith('gerçekleşen') else None
            kolonlar['ger_b'] = (kolonlar['plan_b'] + 1) if kolonlar['plan_b'] and 'unit cost' in norm(hdr.get(kolonlar['plan_b'] + 1, '')) else None
            continue
        if not hdr or not kolonlar.get('mecra'): continue

        mecra = txt(ws.cell(r, kolonlar['mecra']).value)
        format_ = txt(ws.cell(r, kolonlar['format']).value) if kolonlar['format'] else ''
        if not mecra or not format_: continue

        g_grup, g_ad, g_yok = SZ.yayinci(mecra)
        t_bilgi = SZ.yayin_turu(format_)
        plan_h = num(ws.cell(r, kolonlar['plan_h']).value) if kolonlar['plan_h'] else None
        plan_b = num(ws.cell(r, kolonlar['plan_b']).value) if kolonlar['plan_b'] else None
        plan_bd = num(ws.cell(r, kolonlar['plan_bd']).value) if kolonlar['plan_bd'] else None
        ger_bd = num(ws.cell(r, kolonlar['ger_bd']).value) if kolonlar['ger_bd'] else None
        ger_b = num(ws.cell(r, kolonlar['ger_b']).value) if kolonlar.get('ger_b') else None

        tip, tip_kaynak = SZ.fiyat_tipi_bul(plan_b, plan_h, plan_bd, format_)
        if ger_b is None:
            ozel = {'CPM': 'ger_cpm', 'CPV': 'ger_cpv'}.get(tip)
            if ozel and kolonlar.get(ozel):
                ger_b = num(ws.cell(r, kolonlar[ozel]).value)

        if kolonlar.get('ger_h'):
            ger_h = num(ws.cell(r, kolonlar['ger_h']).value)
        else:  # Züber tipi: hacmi fiyat tipine göre seç
            sec = {'CPM': 'imp', 'CPV': 'view', 'CPC': 'click'}.get(tip, 'imp')
            ger_h = num(ws.cell(r, kolonlar[sec]).value) if kolonlar.get(sec) else None
            if ger_h is None and kolonlar.get('imp'): ger_h = num(ws.cell(r, kolonlar['imp']).value)

        tik = num(ws.cell(r, kolonlar['click']).value) if kolonlar.get('click') else None
        vcr = num(ws.cell(r, kolonlar['vcr']).value) if kolonlar.get('vcr') else None
        if vcr is not None and vcr > 1.5: vcr = vcr / 100 if vcr <= 100 else None
        ahb = num(ws.cell(r, kolonlar['ahb']).value) if kolonlar.get('ahb') else None
        adsrv = num(ws.cell(r, kolonlar['adsrv']).value) if kolonlar.get('adsrv') else None
        satir_tarih = txt(ws.cell(r, kolonlar['tarih']).value) if kolonlar.get('tarih') else ''

        if ger_h and ger_h > 0:
            durum = 'Yayınlandı'
        elif (plan_bd or 0) > 0:
            durum = 'Veri Gelmedi'; uyari.append(f'satır {r}: {mecra} — gerçekleşen rakam yok')
        else:
            durum = 'Planlandı'

        satirlar.append(dict(
            kampanya_kod=None, marka=marka, kampanya=kampanya, s1=s1, s2=s2, s3=s3,
            bas=b_tar, bit=s_tar, tarih_kesin=kesin, satir_tarih=satir_tarih,
            amac=t_bilgi['amac'],
            grup=g_grup, mecra=g_ad, mecra_ham=mecra, sozlukte_yok=(g_yok or t_bilgi['sozlukte_yok']),
            ana_tur=t_bilgi['ana'], reklam_modeli=(t_bilgi['reklam_modeli'] or t_bilgi['ana']),
            site=txt(ws.cell(r, kolonlar['site']).value) if kolonlar['site'] else '',
            hedef=txt(ws.cell(r, kolonlar['hedef']).value) if kolonlar['hedef'] else '',
            cihaz=txt(ws.cell(r, kolonlar['cihaz']).value) if kolonlar['cihaz'] else '',
            format=format_, aset=txt(ws.cell(r, kolonlar['aset']).value) if kolonlar['aset'] else '',
            tip=tip, tip_kaynak=tip_kaynak,
            plan_birim=plan_b, plan_hacim=plan_h, plan_bedel=plan_bd,
            ger_hacim=ger_h, ger_birim=ger_b, ger_bedel=ger_bd,
            tik=tik, vcr=vcr, ahb=ahb, adserver=adsrv,
            kaynak=kaynak(plan_b, ger_b), durum=durum, kaynak_dosya=ad))
    return dict(dosya=ad, marka=marka, sektor_yok=sektor_yok, kampanya=kampanya, donem=donem, yil=yil,
                bas=b_tar, bit=s_tar, tarih_kesin=kesin, s1=s1, s2=s2, s3=s3,
                butce=butce, satirlar=satirlar, uyari=uyari)

def calistir(klasor, sozluk_yolu=None):
    global SZ
    if sozluk_yolu is None:
        sozluk_yolu = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                   '..', 'sozluk', 'sozluk.xlsx')
    SZ = Sozluk(sozluk_yolu)
    sonuc = []
    for f in sorted(os.listdir(klasor)):
        if not f.endswith('.xlsx') or f.startswith('~'): continue
        try:
            r = oku_dosya(os.path.join(klasor, f))
        except Exception as e:
            print('HATA', f, e); continue
        if r: sonuc.append(r); print(f'{f}: {len(r["satirlar"])} satır')
        else: print(f'{f}: tanınmadı')
    print(); print(SZ.rapor())
    return sonuc


if __name__ == '__main__':
    klasor = sys.argv[1] if len(sys.argv) > 1 else '.'
    sz_yol = sys.argv[2] if len(sys.argv) > 2 else None
    sonuc = calistir(klasor, sz_yol)
    with open('/tmp/donusum.json', 'w', encoding='utf-8') as fh:
        json.dump(sonuc, fh, ensure_ascii=False, default=str)
