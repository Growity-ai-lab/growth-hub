# -*- coding: utf-8 -*-
"""Dönüştürülen veriden kampanya başına doldurulmuş şablon üretir. Puanlama bu dosyada YOKTUR (bkz. /puanlama/skor.py)."""
import json, math, os, shutil, statistics, datetime, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KOK_DIZIN = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SABLON = os.path.join(KOK_DIZIN, 'sablon', 'TG_Yayin_Plani_Sablonu_v1.xlsx')
OUT = os.path.join(KOK_DIZIN, 'veri')
KAMP = os.path.join(OUT, 'donusturulmus')
os.makedirs(KAMP, exist_ok=True)
D = json.load(open('/tmp/donusum.json', encoding='utf-8'))
BUGUN = datetime.date(2026, 8, 4)

F = 'Arial'
BLUE = Font(name=F, size=10, color='0000FF')
BLACK = Font(name=F, size=10)
BOLD = Font(name=F, size=10, bold=True)
HDR_FONT = Font(name=F, size=9, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor='1F3864')
GRY = PatternFill('solid', fgColor='F2F2F2')
thin = Side(style='thin', color='BFBFBF'); BOX = Border(thin, thin, thin, thin)

def kod(k):
    m = re.sub(r'[^A-Za-z]', '', (k['marka'] or 'XX'))[:4].upper()
    kk = re.sub(r'[^A-Za-z]', '', (k['kampanya'] or 'XX'))[:4].upper()
    return f"{m}-{kk}-{k['bas'][:7].replace('-','')}"

def aile(satir):
    return satir.get('ana_tur') or 'Diğer'


# ------------------------------------------------------------------ 1) kampanya dosyaları
for k in D:
    k['kod'] = kod(k)
    hedef = os.path.join(KAMP, f"{k['kod']}.xlsx")
    shutil.copy(SABLON, hedef)
    wb = openpyxl.load_workbook(hedef)
    kp, pg = wb['Kampanya'], wb['Plan ve Gerçekleşen']

    for hucre, deger in [('C4', k['kod']), ('C5', k['marka']), ('C6', k['marka']), ('C7', k['kampanya']),
                         ('C8', k['s1']), ('C9', k['s2']), ('C10', k['s3']),
                         ('C11', k['bas']), ('C12', k['bit']), ('C13', 'Kapanış'),
                         ('C14', BUGUN.isoformat())]:
        kp[hucre] = deger
    kp['C23'] = k['butce'] or 0
    kp['C19'] = round(sum(x['adserver'] or 0 for x in k['satirlar']), 2)

    # örnek satırları temizle
    for r in range(3, 53):
        for c in list(range(1, 17)) + [18, 19, 20, 21, 22, 23, 28, 30, 32, 35, 36]:
            pg.cell(r, c).value = None

    for i, x in enumerate(k['satirlar']):
        r = 3 + i
        if r > 52: break
        ahb_oran = (x['ahb'] / x['ger_bedel']) if (x['ahb'] and x['ger_bedel']) else 0
        notlar = []
        if x['kaynak'] == 'Hesaplanan':
            notlar.append('kaynak dosyada gerçekleşen birim maliyet plandan kopyalanmış; fiyat farkı ölçülemez')
        if not k['tarih_kesin']:
            notlar.append('tarih kampanya döneminden türetildi')
        if x['tip_kaynak'] != 'oran hesabı':
            notlar.append(f"fiyat tipi {x['tip_kaynak']} ile tahmin edildi")
        vals = {1: f"S{i+1:03d}", 2: k['kod'], 3: k['bas'], 4: k['bit'], 5: x['amac'],
                6: x['grup'], 7: x['mecra'], 8: x['site'] or x['mecra'], 9: x['site'],
                10: x['cihaz'], 11: x['format'], 12: x['hedef'], 13: None,
                14: x['tip'], 15: x['plan_birim'], 16: x['plan_hacim'],
                18: x['ger_hacim'], 19: x['tik'], 21: None,
                22: ('Platform Raporu' if x['kaynak'] == 'Platform Raporu' else 'Fatura') if x['ger_bedel'] else None,
                23: x['ger_bedel'], 28: round(ahb_oran, 4), 30: 0, 32: 0,
                35: x['durum'], 36: ' · '.join(notlar)}
        for c, v in vals.items():
            cell = pg.cell(r, c); cell.value = v; cell.font = BLUE; cell.border = BOX
    wb.save(hedef)

