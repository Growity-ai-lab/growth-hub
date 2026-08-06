# -*- coding: utf-8 -*-
"""
Sözlük yükleyici — sozluk.xlsx'i okur, diğer tüm modüller bunu kullanır.
Kod içinde hiçbir yayıncı, yayın türü, sektör ya da parametre gömülü değildir.

Kullanım:
    from sozluk import Sozluk
    sz = Sozluk('.../sozluk/sozluk.xlsx')
    sz.yayinci('Youtube')        -> ('Google', 'Youtube', False)   # grup, standart ad, sözlükte yok mu
    sz.yayin_turu('Trivuew')     -> {'ana': 'Kısa video', ...}
    sz.sektor('Uludağ')          -> ('FMCG', 'İçecek', 'Gazlı ve Meyveli İçecek', False)
    sz.bolen('CPM')              -> 1000
    sz.p('tek_grup_tavani')      -> 0.30
"""
import re, unicodedata
import openpyxl


def norm(s):
    if s is None: return ''
    s = str(s).replace('\n', ' ').replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s).strip().lower()
    s = s.replace('i̇', 'i')
    return unicodedata.normalize('NFC', s)


class Sozluk:
    def __init__(self, yol):
        self.yol = yol
        wb = openpyxl.load_workbook(yol, data_only=True)
        self.eksikler = {'yayinci': set(), 'yayin_turu': set(), 'sektor': set()}

        def oku(sayfa):
            ws = wb[sayfa]
            bas = [norm(c.value) for c in ws[1]]
            cikti = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r[0] in (None, ''): continue
                cikti.append(dict(zip(bas, r)))
            return cikti

        self.yayincilar = {}
        for s in oku('Yayıncılar'):
            if norm(s.get('kullanımda mı')) == 'hayır': continue
            self.yayincilar[norm(s['raporda geçen ad'])] = (
                (s.get('yayıncı grubu') or s['raporda geçen ad']).strip(),
                (s.get('standart yayıncı adı') or s['raporda geçen ad']).strip())

        self.turler = {}
        for s in oku('Yayın Türleri'):
            if norm(s.get('kullanımda mı')) == 'hayır': continue
            self.turler[norm(s['raporda geçen ad'])] = {
                'ana': (s.get('ana yayın türü') or 'Diğer').strip(),
                'reklam_modeli': (s.get('standart reklam modeli') or '').strip(),
                'beklenen_tip': (s.get('beklenen fiyat tipi') or 'CPM').strip(),
                'video': norm(s.get('video mu')) == 'evet',
                'amac': (s.get('kampanya amacı') or 'Gösterim almak').strip()}

        self.sektorler = {}
        for s in oku('Sektörler'):
            if norm(s.get('kullanımda mı')) == 'hayır': continue
            self.sektorler[norm(s['marka'])] = (
                (s.get('sektör l1') or '').strip(), (s.get('sektör l2') or '').strip(),
                (s.get('sektör l3') or '').strip())

        self.tipler = {}
        for s in oku('Fiyat Tipleri'):
            if norm(s.get('kullanımda mı')) == 'hayır': continue
            self.tipler[(s['fiyat tipi'] or '').strip()] = float(s.get('bölen') or 1)

        self.parametreler = {}
        for s in oku('Parametreler'):
            self.parametreler[norm(s['parametre'])] = s.get('değer')

    # -------------------------------------------------------------- arama
    def _ara(self, sozluk, ad, kutu):
        k = norm(ad)
        if not k: return None
        if k in sozluk: return sozluk[k]
        for anahtar, deger in sozluk.items():          # kısmi eşleşme
            if anahtar and (k.startswith(anahtar + ' ') or anahtar in k):
                return deger
        self.eksikler[kutu].add(str(ad).strip())
        return None

    def yayinci(self, ad):
        v = self._ara(self.yayincilar, ad, 'yayinci')
        if v: return v[0], v[1], False
        return (str(ad).strip() or 'Bilinmiyor'), str(ad).strip(), True

    def yayin_turu(self, ad):
        v = self._ara(self.turler, ad, 'yayin_turu')
        if v: return dict(v, sozlukte_yok=False)
        return {'ana': 'Diğer', 'reklam_modeli': '', 'beklenen_tip': 'CPM', 'video': False,
                'amac': 'Gösterim almak', 'sozlukte_yok': True}

    def sektor(self, marka):
        v = self._ara(self.sektorler, marka, 'sektor')
        if v: return v[0], v[1], v[2], False
        return 'Bilinmiyor', '', '', True

    def bolen(self, tip):
        return self.tipler.get(tip, 1.0)

    def p(self, ad, varsayilan=None):
        v = self.parametreler.get(norm(ad), varsayilan)
        if isinstance(v, str) and '|' in v:
            return [x.strip() for x in v.split('|')]
        return v

    # -------------------------------------------------------------- fiyat tipi
    def fiyat_tipi_bul(self, plan_birim, plan_hacim, plan_bedel, yayin_turu_adi):
        """Eski raporlarda fiyat tipi yazmaz; tutar/(fiyat×adet) oranından bulunur."""
        bilgi = self.yayin_turu(yayin_turu_adi)
        beklenen = bilgi['beklenen_tip']
        tikla = bilgi['amac'] in ('Siteye trafik', 'Satış yaptırmak')
        birim_basi = beklenen if self.bolen(beklenen) == 1 else ('CPC' if tikla else 'CPV')
        if plan_birim and plan_hacim and plan_bedel:
            oran = plan_bedel / (plan_birim * plan_hacim)
            if 0.0006 < oran < 0.0016: return 'CPM', 'oran hesabı'
            if 0.5 < oran < 1.6:
                return birim_basi, 'oran hesabı'
        if plan_birim and plan_birim >= 5: return 'CPM', 'fiyat büyüklüğü'
        return beklenen, 'sözlükteki beklenen tip'

    def rapor(self):
        satir = [f'Sözlük: {self.yol}',
                 f"  yayıncı {len(self.yayincilar)} · yayın türü {len(self.turler)} · "
                 f"marka {len(self.sektorler)} · fiyat tipi {len(self.tipler)} · "
                 f"parametre {len(self.parametreler)}"]
        for kutu, ad in [('yayinci', 'Yayıncılar'), ('yayin_turu', 'Yayın Türleri'), ('sektor', 'Sektörler')]:
            if self.eksikler[kutu]:
                satir.append(f'  SÖZLÜKTE YOK — {ad} sayfasına eklenmeli: '
                             + ', '.join(sorted(self.eksikler[kutu])))
        return '\n'.join(satir)
